from openpyxl import workbook
from openpyxl.xml.constants import WORKBOOK_MACRO
import pandas as pd
import pprint
import openpyxl
import numpy
from datetime import datetime
import re

def write_dict_to_excel(file="",content=[]):
    try:
        workbook = openpyxl.load_workbook(file)
        for sheet in content:
            sheet_name = sheet['sheet_name']
            if not("col_list" in sheet.keys()):
                now = datetime.now().strftime('%d/%m/%Y:%H:%M')
                print(f"{now} : could not write into file {file} sheet: {sheet['sheet_name']}")
                continue
            for items in sorted(workbook[sheet_name].merged_cell_ranges):
                workbook[sheet_name].unmerge_cells(str(items))
            for row in range(1,workbook[sheet_name].max_row+1):
                for col in range(1,workbook[sheet_name].max_column+1):
                    workbook[sheet_name].cell(row,col).value = None
            
            # Fill the first row of the Excel sheet
            valid_index = 1
            for index in range(0,len(sheet['col_list'])):
                if re.match(r'.*Unnamed.*', sheet['col_list'][index]):
                    #workbook[sheet_name].cell(1,index+1).value = sheet['col_list'][index]
                    workbook[sheet_name].merge_cells(start_row=1,start_column=valid_index,end_row=1,end_column=index+1)
                else:
                    valid_index=index+1
                    workbook[sheet_name].cell(1,index+1).value = sheet['col_list'][index]
            row=2
            for content in sheet['content']:
                for key in content.keys():
                    col = sheet['col_list'].index(key) + 1
                    workbook[sheet_name].cell(row,col).value = content[key]
                row+=1

        workbook.save(file)

        # writer = pd.ExcelWriter(file, engine='xlsxwriter')
        # for frame in content:
        #     df = pd.DataFrame(frame["content"])
        #     df.to_excel(writer,sheet_name=frame["sheet_name"],index=False)
        # writer.save()
    except:
        now = datetime.now().strftime('%d/%m/%Y:%H:%M')
        print(f"{now} : {file} was not written")

def read_dict_from_excel(file=""):

    try:
        workbook = openpyxl.load_workbook(file)
        dict = []
        for sheet in workbook.sheetnames:
            col_list = []
            sheet_dict = {}
            sheet_dict.update({"sheet_name" : sheet})
            for col in range(1,workbook[sheet].max_column+1):
                value = workbook[sheet].cell(1,col).value
                if value is not None:
                    col_list.append(value)
                else:
                    col_list.append(f"Unnamed: {col-1}")
            sheet_dict.update({"col_list" : col_list})
            content = []
            for row in range(2,workbook[sheet].max_row+1):
                row_dict = {}
                for col in range(1,workbook[sheet].max_column+1):
                    value = workbook[sheet].cell(row,col).value
                    if value is None:
                        value = numpy.nan
                    row_dict.update({col_list[col-1]:value})
                content.append(row_dict)
            sheet_dict.update({"content":content})
            dict.append(sheet_dict)
        return(dict)

        # df = pd.read_excel(file,sheet_name=None)
        # dict = []
        # for key in df.keys():
        #     sheet_content = df[key].to_dict('records')
        #     dict.append({"sheet_name" : key,
        #                  "content" : sheet_content})

    except:
        print("Failed")

