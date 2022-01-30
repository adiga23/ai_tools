from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service, service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from datetime import date, datetime
import json
import os
import keyring
from pprint import pprint
import time
import re
from openpyxl import workbook
from openpyxl.xml.constants import WORKBOOK_MACRO
import pandas as pd
import openpyxl
import numpy

def write_dict_to_excel(file="",content=[]):

    workbook = openpyxl.load_workbook(file)
    for sheet in content:
        sheet_name = sheet['sheet_name']
        if not("col_list" in sheet.keys()):
            now = datetime.now().strftime('%d/%m/%Y:%H:%M')
            print(f"{now} : could not write into file {file} sheet: {sheet['sheet_name']}")
            continue
        for items in sorted(workbook[sheet_name].merged_cell_ranges):
            workbook[sheet_name].unmerge_cells(str(items))
        for row in range(1,workbook[sheet_name].max_row+1):
            for col in range(1,workbook[sheet_name].max_column+1):
                workbook[sheet_name].cell(row,col).value = None

        # Fill the first row of the Excel sheet
        valid_index = 1
        for index in range(0,len(sheet['col_list'])):
            if re.match(r'.*Unnamed.*', sheet['col_list'][index]):
                #workbook[sheet_name].cell(1,index+1).value = sheet['col_list'][index]
                workbook[sheet_name].merge_cells(start_row=1,start_column=valid_index,end_row=1,end_column=index+1)
            else:
                valid_index=index+1
                workbook[sheet_name].cell(1,index+1).value = sheet['col_list'][index]
        row=2
        for content in sheet['content']:
            for key in content.keys():
                col = sheet['col_list'].index(key) + 1
                workbook[sheet_name].cell(row,col).value = content[key]
            row+=1

    workbook.save(file)

def read_dict_from_excel(file=""):

    try:
        workbook = openpyxl.load_workbook(file)
        dict = []
        for sheet in workbook.sheetnames:
            col_list = []
            sheet_dict = {}
            sheet_dict.update({"sheet_name" : sheet})
            for col in range(1,workbook[sheet].max_column+1):
                value = workbook[sheet].cell(1,col).value
                if value is not None:
                    col_list.append(value)
                else:
                    col_list.append(f"Unnamed: {col-1}")
            sheet_dict.update({"col_list" : col_list})
            content = []
            for row in range(2,workbook[sheet].max_row+1):
                row_dict = {}
                for col in range(1,workbook[sheet].max_column+1):
                    value = workbook[sheet].cell(row,col).value
                    if value is None:
                        value = numpy.nan
                    row_dict.update({col_list[col-1]:value})
                content.append(row_dict)
            sheet_dict.update({"content":content})
            dict.append(sheet_dict)
        return(dict)

        # df = pd.read_excel(file,sheet_name=None)
        # dict = []
        # for key in df.keys():
        #     sheet_content = df[key].to_dict('records')
        #     dict.append({"sheet_name" : key,
        #                  "content" : sheet_content})

    except:
        print("Failed")

def element_click_send_key(driver,id,data):
    element = WebDriverWait(driver, 30).until(
              EC.element_to_be_clickable((By.XPATH, id))
              )
    element.send_keys(data)
    return

def element_view(driver,id):
    element = driver.find_element(By.XPATH,id)
    driver.execute_script("arguments[0].scrollIntoView();",element)

def element_click(driver,id):
    element = WebDriverWait(driver, 30).until(
              EC.element_to_be_clickable((By.XPATH, id))
              )
    element.click()
    return

def element_clikable(element):
    temp_element = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable(element)
                   )
    return

def get_next_page(driver):
    pages = driver.find_elements(By.XPATH,"//li[@class='pagination--li']")
    num_pages = len(pages)
    try:
        next_page_link=pages[num_pages-1].find_element(By.XPATH,".//a").get_attribute('href')
    except:
        next_page_link=None
    return(next_page_link)

def get_car_dict(car_type,status_dict):
    existing_ids = {}
    mod_status_dict = []
    content = []
    content_new = []
    for sheet in status_dict:
        if car_type == sheet['sheet_name']:
            mod_status_dict.append({'sheet_name' : car_type,
                                    'col_list' : sheet['col_list'],
                                    'content' : []})
            for entry in sheet['content']:
                id_link = entry['Link']
                id = re.sub(r'.*?,(.*?)\)',r'\1',id_link)
                id = re.sub(r'.*?\"(.*)\"',r'\1',id)
                existing_ids.update({f"{id}" : {'Price'     : entry['Price'],
                                                'Year'      : entry['Year'],
                                                'Car make'  : entry['Car make'],
                                                'Car Model' : entry['Car Model'],
                                                'Distance'  : entry['Distance'],
                                                'Comments'  : entry['Comments'] }})
        else:
            mod_status_dict.append(sheet)


    try:
        exec_path = "/Users/raga/python_scripts/geckodriver"
        firefox_option = Options()
        firefox_option.add_argument("--window-size=1920,1080")
        firefox_option.add_argument("--incognito")
        firefox_option.add_argument("--headless")
        s=Service(executable_path=exec_path,log_path=os.devnull)
        driver = webdriver.Firefox(options=firefox_option,service=s)

        driver.get(f"https://www.autotrader.co.uk/car-search?postcode=CB19YG&make={car_type}&price-from=12000&price-to=20000&include-delivery-option=on&body-type=SUV&maximum-mileage=50000&transmission=Automatic&year-from=2015&minimum-badge-engine-size=2.0&min-engine-power=150&ulez-compliant=on&exclude-writeoff-categories=on&advertising-location=at_cars&keywords=cruise&keywords=leather")
        try:
            element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, ".//button[text()='Accept All']"))
            )  
            element.click()
        except:
            pass

        driver.get(f"https://www.autotrader.co.uk/car-search?postcode=CB19YG&make={car_type}&price-from=12000&price-to=20000&include-delivery-option=on&body-type=SUV&maximum-mileage=50000&transmission=Automatic&year-from=2015&minimum-badge-engine-size=2.0&min-engine-power=150&ulez-compliant=on&exclude-writeoff-categories=on&advertising-location=at_cars&keywords=cruise&keywords=leather")

       
        ## Get all the list of the cars as keys which is a id from which
        ## we can form a link to get the car
        car_list={}
        while True:
            page_list = driver.find_elements(By.XPATH,"//li[@class='search-page__result']")
            for car in page_list:
                id = car.get_attribute('id')
                distance = car.get_attribute('data-distance-value')
                car_list.update({f"{id}" : {"link" : f"https://www.autotrader.co.uk/car-details/{id}",
                                        "distance" : distance}})
            next_page = get_next_page(driver)
            if next_page==None:
                break
            else:
                driver.get(next_page)

        for id in car_list.keys():
            print(f"Getting data for {id}")
            time.sleep(2)
            driver.get(car_list[id]["link"])
            element = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//h2[@data-gui='advert-price']"))
                    )
            price = re.sub(r'.*?(\d+)',r'\1',element.text)
            try:
                mileage = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//span[@data-gui='mileage']"))
                        )
                #mileage = mileage.text
                print(f"I am here {mileage.text}")
            except:
                mileage = "0 miles"

            try:
                car_model = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//h1[contains(@class,'sc-hUpaCq bNxgLI')]"))
                        )
                car_model = car_model.text
            except:
                car_model = "Unknown"
            try:
                registration = WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.XPATH, "//p[contains(@class,'sc-jgrJph khASqk')]"))
                            )
                registration = registration.text
            except:
                car_model = "Unknown"
            try:
                car_make = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.XPATH, "//p[contains(@class,'sc-gSQFLo kmZccV')]"))
                        )
                car_make = car_make.text
            except:
                car_make = "Unknown"
            try:
                market_cmp=driver.find_element(By.XPATH,"//div[contains(@class,'atc-type-smart--medium')]").text
            except:
                market_cmp = ""
            car_details={'Link'      : f'=HYPERLINK("https://www.autotrader.co.uk/car-details/{id}","{id}")',
                        'Price'     : price,
                        'Year'      : registration,
                        'Car make'  : car_make,
                        'Car Model' : car_model,
                        'Distance'  : car_list[id]["distance"],
                        'Price cmp' : market_cmp,
                        'mileage'   : mileage,
                        'Comments'  : ""}
            car_details_new={'Link'      : f'=HYPERLINK("https://www.autotrader.co.uk/car-details/{id}","{id}")',
                            'Price'     : price,
                            'Year'      : registration,
                            'Car make'  : car_make,
                            'Car Model' : car_model,
                            'Distance'  : car_list[id]["distance"],
                            'Price cmp' : market_cmp,
                            'mileage'   : mileage,
                            'Comments'  : ""}
            if id in existing_ids.keys():
                car_details['Comments'] = existing_ids[id]['Comments']
                content.append(car_details)
                if price != existing_ids[id]['Price']:
                    car_details['Comments'] = f"Price change from {existing_ids[id]['Price']} to {price}"
                    content_new.append(car_details)
            else:
                content.append(car_details)
                car_details_new['Comments'] = "New Car"
                content_new.append(car_details_new)
    except:
        print(f"Issue getting details for car {car_type}")
    try:
        driver.quit()
    except:
        pass

    for sheet in mod_status_dict:
        if sheet['sheet_name'] == "NEW":
            sheet['content'] += content_new
        elif sheet['sheet_name'] == car_type:
            sheet['content'] = content
    return(mod_status_dict)

status_file = "car_prices.xlsx"

status_dict = read_dict_from_excel(status_file)

# Empty the contents in the new sheet
for sheet in status_dict:
    if sheet["sheet_name"] == "NEW":
        sheet["content"] = []

now=datetime.now().strftime("%d/%m/%Y")

f = open('timestamp','r')
lines = f.readlines()
f.close()

trigger_parse = (len(lines) == 0) or \
                (now != lines[0])

mod_status_dict = get_car_dict('Audi',status_dict)
pprint(mod_status_dict)

trigger_parse = False
if trigger_parse:
    error_found = False
    print("Getting details for Audi")
    try:
        mod_status_dict = get_car_dict('Audi',status_dict)
        if mod_status_dict == None:
            error_found = True
        else:
            status_dict = mod_status_dict
    except:
        error_found = True

    print("Getting details for BMW")
    try:
        mod_status_dict = get_car_dict('BMW',status_dict)
        if mod_status_dict == None:
            error_found = True
        else:
            status_dict = mod_status_dict
    except:
        error_found = True

    print("Getting details for Benz")
    try:
        mod_status_dict = get_car_dict('Mercedes-Benz',status_dict)
        if mod_status_dict == None:
            error_found = True
        else:
            status_dict = mod_status_dict
    except:
        error_found = True

    try:
        write_dict_to_excel(status_file,status_dict)
    except:
        error_found = True

    if not error_found:
        f = open('timestamp','w')
        f.write(now)
        f.close()
else:
    print("Not updating car prices")
